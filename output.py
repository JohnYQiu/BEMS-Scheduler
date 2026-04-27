"""
output.py
=========
Handles all output for the Brown EMS scheduler:
  - Formatted .xlsx file with multiple sheets:
      1. Schedule         — week-by-week ambulance grid (Sun → Sat)
      2. Campus Response  — week-by-week campus responder grid (Sun → Sat)
      3. Hour Summary     — totals (ambulance + campus)
      4. Warnings         — unfilled shifts, ALS without EVDT, no auth driver (night/weekend only)
      5. Strike List      — missing minimum availability categories
  - Brief terminal summary
"""

from datetime import date, timedelta
from scheduler import Shift
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SHIFT_TIMES = {
    "AM":    ("0700", "1300"),
    "PM":    ("1300", "1900"),
    "NIGHT": ("1900", "0700+1"),
    "DAY":   ("0700", "1900"),
}

C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_DATE_BG   = "2E5FA3"
C_DATE_FG   = "FFFFFF"
C_ALS_BG    = "FCE4D6"
C_EVDT_BG   = "E2EFDA"
C_AUTH_BG   = "FFF2CC"
C_EMT_BG    = "FFFFFF"
C_UNFILLED  = "F4CCCC"
C_WARN_BG   = "FFE0E0"
C_ALT_ROW   = "F8F8F8"
C_UNDER18   = "CC0000"

CERT_BG = {"EVDT": C_EVDT_BG, "Auth": C_AUTH_BG, "EMT": C_EMT_BG}
C_BERT_BG = "D9E1F2"

_thin   = Side(style="thin", color="CCCCCC")
BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, values, widths=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=True, color=C_HEADER_FG)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align(h="center")
        c.border    = BORDER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _week_start_sunday(d: date) -> date:
    # Python weekday: Mon=0 ... Sun=6
    return d - timedelta(days=(d.weekday() + 1) % 7)


def _warn_no_auth_driver(shift: Shift) -> bool:
    """Weekday AM/PM crews do not require a separate auth-driver warning."""
    if shift.date.weekday() < 5 and shift.shift_type in ("AM", "PM"):
        return False
    return not shift.has_auth


def _slots_for_ambulance_shift(shift: Shift) -> dict:
    """
    Return role buckets:
      - evdts: list of all EVDTs
      - auths: list of Auth-only drivers (not EVDT)
      - emts:  list of EMT-only (not Auth-capable)
    """
    evdts = [v for v in shift.volunteers if v.is_evdt]
    auths = [v for v in shift.volunteers if v.certification == "Auth"]
    emts = [v for v in shift.volunteers if not v.is_auth]
    return {"evdts": evdts, "auths": auths, "emts": emts}


def _build_schedule_sheet(ws, all_shifts):
    ws.title = "Schedule"
    ws.freeze_panes = "C3"

    # Columns:
    #   A: Shift label (merged over 4 rows)
    #   B: Slot label (EVDT/Auth/EMT1/EMT2)
    #   C..I: Sun..Sat
    widths = [10, 10, 18, 18, 18, 18, 18, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    if not all_shifts:
        return

    all_dates = sorted({d for (d, _) in all_shifts.keys()})
    start = _week_start_sunday(all_dates[0])
    end = all_dates[-1]

    day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    shift_sections = [("DAY", "0700-1900"), ("AM", "0700-1300"), ("PM", "1300-1900"), ("NIGHT", "1900-0700+1")]
    slot_rows = [("EVDT", "evdt"), ("Auth", "auth"), ("EMT1", "emt1"), ("EMT2", "emt2")]

    row = 1
    current = start
    while current <= end:
        week_label = f"Week of {current.isoformat()} (Sun–Sat)"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=week_label)
        c.font = _font(bold=True, color=C_DATE_FG)
        c.fill = _fill(C_DATE_BG)
        c.alignment = _align(h="left")
        c.border = BORDER
        row += 1

        headers = ["Shift", "Slot"] + [f"{day_names[i]} { (current + timedelta(days=i)).strftime('%m/%d') }" for i in range(7)]
        _header_row(ws, row, headers)
        row += 1

        for shift_type, time_str in shift_sections:
            start_row = row
            for slot_label, slot_key in slot_rows:
                # Shift label / slot label columns
                ws.cell(row=row, column=2, value=slot_label).border = BORDER
                ws.cell(row=row, column=2).alignment = _align(h="center")
                ws.cell(row=row, column=2).font = _font(bold=True)

                for i in range(7):
                    d = current + timedelta(days=i)
                    key = (d, shift_type)
                    shift = all_shifts.get(key)
                    val = ""
                    fill = _fill(C_EMT_BG)

                    if shift:
                        slots = _slots_for_ambulance_shift(shift)
                        if slot_key == "evdt":
                            evdts = slots["evdts"]
                            if evdts:
                                val = "\n".join(v.full_name for v in evdts)
                                fill = _fill(C_EVDT_BG)
                        elif slot_key == "auth":
                            auths = slots["auths"]
                            if auths:
                                val = "\n".join(v.full_name for v in auths)
                                fill = _fill(C_AUTH_BG)
                        else:
                            emts = slots["emts"]
                            idx = 0 if slot_key == "emt1" else 1
                            if idx < len(emts):
                                v = emts[idx]
                                val = v.full_name
                                fill = _fill(CERT_BG.get(v.certification, C_EMT_BG))

                    cell = ws.cell(row=row, column=3 + i, value=val)
                    cell.fill = fill
                    cell.border = BORDER
                    cell.alignment = _align(h="left", wrap=True)
                    cell.font = _font(size=9, bold=bool(val))

                row += 1

            # Merge shift label
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            c = ws.cell(row=start_row, column=1, value=f"{shift_type}\n{time_str}")
            c.alignment = _align(h="center", wrap=True)
            c.font = _font(bold=True)
            c.fill = _fill(C_ALT_ROW)
            c.border = BORDER
            for r in range(start_row + 1, row):
                ws.cell(row=r, column=1).border = BORDER
                ws.cell(row=r, column=1).fill = _fill(C_ALT_ROW)

            # Spacer
            row += 1

        # Bigger spacer between weeks
        row += 1
        current += timedelta(days=7)


def _format_ambulance_shift_list(v) -> str:
    """Human-readable list of ambulance assignments for hour-summary audit."""
    keys = getattr(v, "scheduled_shifts", None) or []
    if not keys:
        return ""
    lines = []
    for d, s in sorted(keys, key=lambda k: (k[0], k[1])):
        lines.append(f"{d.month}/{d.day} {s}")
    return "; ".join(lines)

def _format_campus_shift_list(v) -> str:
    keys = getattr(v, "campus_scheduled_shifts", None) or []
    if not keys:
        return ""
    lines = []
    for d, b in sorted(keys, key=lambda k: (k[0], k[1])):
        lines.append(f"{d.month}/{d.day} {b}")
    return "; ".join(lines)


def _build_summary_sheet(
    ws,
    volunteers,
    ambulance_target_hours: int = 18,
    campus_emt_max_hours: int = 6,
    campus_bert_max_hours: int = 9,
):
    ws.title = "Hour Summary"
    ws.freeze_panes = "A2"
    headers = [
        "Name",
        "Email",
        "Role",
        "Certification",
        "Ambulance Hours",
        "Ambulance shifts",
        "Campus Hours",
        "Campus shifts",
        "Ambulance Status",
        "Campus Target",
    ]
    widths = [28, 32, 10, 14, 12, 40, 12, 32, 18, 18]
    _header_row(ws, 1, headers, widths)

    sorted_people = sorted(volunteers, key=lambda v: (-(getattr(v, "scheduled_hours", 0)), -(getattr(v, "campus_scheduled_hours", 0))))
    for i, v in enumerate(sorted_people, 2):
        amb = getattr(v, "scheduled_hours", 0)
        campus_hours = getattr(v, "campus_scheduled_hours", 0)
        cert = getattr(v, "certification", "")
        is_bert = cert == "BERT"
        role = "BERT" if is_bert else "AMB"
        amb_under = (not is_bert) and amb < ambulance_target_hours
        campus_target = campus_bert_max_hours if is_bert else campus_emt_max_hours
        amb_detail = "—" if is_bert else _format_ambulance_shift_list(v)
        campus_detail = _format_campus_shift_list(v)
        amb_status = "—" if is_bert else (f"⚠ Under {ambulance_target_hours}h" if amb_under else "OK")
        bg    = C_ALT_ROW if i % 2 == 0 else C_EMT_BG
        vals  = [
            v.full_name,
            v.email,
            role,
            cert,
            amb,
            amb_detail,
            campus_hours,
            campus_detail,
            amb_status,
            f"≤ {campus_target}h (ok if under)",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(
                bold=(col == 9 and amb_under),
                color=(C_UNDER18 if (col == 9 and amb_under) else "000000")
            )
            c.alignment = _align(
                h="center" if col in (3, 4, 5, 7, 9, 10) else "left",
                wrap=(col in (6, 8)),
            )
            c.border    = BORDER


def _build_warnings_sheet(ws, all_shifts):
    ws.title = "Warnings"
    headers = ["Type", "Date", "Day", "Shift", "Details"]
    widths  = [25, 13, 12, 8, 45]
    _header_row(ws, 1, headers, widths)

    issues = []
    for key in sorted(all_shifts):
        shift = all_shifts[key]
        if not shift.volunteers:
            issues.append(("UNFILLED SHIFT", shift.date, shift.shift_type, "No volunteers assigned"))
        if shift.has_als and shift.volunteers and not shift.has_evdt:
            names = ", ".join(v.full_name for v in shift.volunteers)
            issues.append(("ALS — NO EVDT", shift.date, shift.shift_type, f"Assigned: {names}"))
        if _warn_no_auth_driver(shift):
            issues.append(("NO AUTH DRIVER", shift.date, shift.shift_type, "No EVDT or Auth on shift"))

    if not issues:
        c = ws.cell(row=2, column=1, value="No warnings — all shifts adequately staffed.")
        c.font = _font(bold=True, color="1E7E34")
        return

    for i, (type_, d, shift_type, detail) in enumerate(issues, 2):
        bg   = "FFF0F0" if i % 2 == 0 else C_WARN_BG
        vals = [type_, d.isoformat(), d.strftime("%A"), shift_type, detail]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(col == 1), color=("CC0000" if col == 1 else "000000"))
            c.alignment = _align(h="center" if col in (2,3,4) else "left")
            c.border    = BORDER


def _build_strike_list_sheet(ws, violations):
    ws.title = "Strike List"
    ws.freeze_panes = "A2"
    headers = ["Name", "Email", "Certification", "Missing Requirements"]
    widths  = [28, 32, 14, 40]
    _header_row(ws, 1, headers, widths)

    if not violations:
        c = ws.cell(row=2, column=1, value="✓ All volunteers met the minimum availability requirements.")
        c.font = _font(bold=True, color="1E7E34")
        return

    for i, item in enumerate(violations, 2):
        v = item["volunteer"]
        missing = ", ".join(item.get("missing", []))
        bg = C_ALT_ROW if i % 2 == 0 else C_EMT_BG
        vals = [v.full_name, v.email, v.certification, missing]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(col == 1))
            c.alignment = _align(h="center" if col in (3,) else "left", wrap=(col == 4))
            c.border    = BORDER


def _build_campus_sheet(ws, campus_shifts):
    ws.title = "Campus Response"
    ws.freeze_panes = "C3"

    widths = [10, 12, 18, 18, 18, 18, 18, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    if not campus_shifts:
        return

    all_dates = sorted({d for (d, _) in campus_shifts.keys()})
    start = _week_start_sunday(all_dates[0])
    end = all_dates[-1]

    day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    blocks = [("A", "0700-1000"), ("B", "1000-1300"), ("C", "1300-1600"), ("D", "1600-1900")]
    slot_rows = [("Responder1", 0), ("Responder2", 1)]

    row = 1
    current = start
    while current <= end:
        week_label = f"Week of {current.isoformat()} (Sun–Sat)"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=week_label)
        c.font = _font(bold=True, color=C_DATE_FG)
        c.fill = _fill(C_DATE_BG)
        c.alignment = _align(h="left")
        c.border = BORDER
        row += 1

        headers = ["Block", "Slot"] + [f"{day_names[i]} { (current + timedelta(days=i)).strftime('%m/%d') }" for i in range(7)]
        _header_row(ws, row, headers)
        row += 1

        for block, time_str in blocks:
            start_row = row
            for slot_label, idx in slot_rows:
                ws.cell(row=row, column=2, value=slot_label).border = BORDER
                ws.cell(row=row, column=2).alignment = _align(h="center")
                ws.cell(row=row, column=2).font = _font(bold=True)

                for i in range(7):
                    d = current + timedelta(days=i)
                    key = (d, block)
                    shift = campus_shifts.get(key)
                    val = ""
                    fill = _fill(C_EMT_BG)

                    if shift and idx < len(getattr(shift, "responders", [])):
                        p = shift.responders[idx]
                        val = getattr(p, "full_name", "")
                        cert = getattr(p, "certification", "BERT")
                        fill = _fill(CERT_BG.get(cert, C_BERT_BG))

                    cell = ws.cell(row=row, column=3 + i, value=val)
                    cell.fill = fill
                    cell.border = BORDER
                    cell.alignment = _align(h="left", wrap=True)
                    cell.font = _font(size=9, bold=bool(val))

                row += 1

            ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            c = ws.cell(row=start_row, column=1, value=f"{block}\n{time_str}")
            c.alignment = _align(h="center", wrap=True)
            c.font = _font(bold=True)
            c.fill = _fill(C_ALT_ROW)
            c.border = BORDER
            for r in range(start_row + 1, row):
                ws.cell(row=r, column=1).border = BORDER
                ws.cell(row=r, column=1).fill = _fill(C_ALT_ROW)

            row += 1

        row += 1
        current += timedelta(days=7)


def export_schedule_xlsx(
    all_shifts,
    volunteers,
    output_path,
    violations=None,
    campus_shifts=None,
    ambulance_target_hours: int = 18,
    campus_emt_max_hours: int = 6,
    campus_bert_max_hours: int = 9,
):
    if not output_path.endswith(".xlsx"):
        output_path = output_path.rsplit(".", 1)[0] + ".xlsx"

    wb = Workbook()
    _build_schedule_sheet(wb.active, all_shifts)
    _build_campus_sheet(wb.create_sheet(), campus_shifts or {})
    _build_summary_sheet(
        wb.create_sheet(),
        volunteers,
        ambulance_target_hours=ambulance_target_hours,
        campus_emt_max_hours=campus_emt_max_hours,
        campus_bert_max_hours=campus_bert_max_hours,
    )
    _build_warnings_sheet(wb.create_sheet(), all_shifts)
    _build_strike_list_sheet(wb.create_sheet(), violations or [])
    wb.save(output_path)
    print(f"  Schedule exported -> {output_path}")
    return output_path


def print_summary(all_shifts, volunteers, ambulance_target_hours: int = 18):
    total       = len(all_shifts)
    unfilled    = sum(1 for s in all_shifts.values() if not s.volunteers)
    als_no_evdt = sum(
        1 for s in all_shifts.values()
        if s.has_als and s.volunteers and not s.has_evdt
    )
    no_auth     = sum(1 for s in all_shifts.values() if _warn_no_auth_driver(s))
    under_tgt   = sum(1 for v in volunteers if v.scheduled_hours < ambulance_target_hours)

    print("\n" + "=" * 55)
    print("SCHEDULE SUMMARY")
    print("=" * 55)
    print(f"  Total shifts:          {total}")
    print(f"  Unfilled shifts:       {unfilled}" + (" ⚠" if unfilled else " ✓"))
    print(f"  ALS shifts w/o EVDT:   {als_no_evdt}" + (" ⚠" if als_no_evdt else " ✓"))
    print(f"  Night/wknd w/o auth:   {no_auth}" + (" ⚠" if no_auth else " ✓"))
    print(
        f"  Volunteers under {ambulance_target_hours}h ambulance:  {under_tgt}"
        + (" ⚠" if under_tgt else " ✓")
    )
    print()


def print_warnings(all_shifts):
    issues = []
    for key in sorted(all_shifts):
        shift = all_shifts[key]
        if not shift.volunteers:
            issues.append(f"UNFILLED:       {shift.label}")
        if shift.has_als and shift.volunteers and not shift.has_evdt:
            issues.append(f"ALS / NO EVDT:  {shift.label}")
        if _warn_no_auth_driver(shift):
            issues.append(f"NO AUTH DRIVER: {shift.label}")

    print("=" * 55)
    print("WARNINGS")
    print("=" * 55)
    if not issues:
        print("  No warnings.")
    else:
        for issue in issues:
            print(f"  ⚠  {issue}")
    print()